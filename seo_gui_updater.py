#!/usr/bin/env python3
"""
Multi-Site WordPress SEO Meta Updater — GUI Version
Works with: Yoast SEO | Rank Math | All in One SEO
Requirements: pip install requests openpyxl
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading
import csv
import json
import base64
import time
import os
import sys
from datetime import datetime

try:
    import requests
    HAS_REQUESTS = True
except ImportError:
    HAS_REQUESTS = False

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ─────────────────────────────────────────────
#  SEO PLUGIN FIELD MAPPINGS
# ─────────────────────────────────────────────
SEO_FIELDS = {
    "yoast": {
        "title": "_yoast_wpseo_title",
        "description": "_yoast_wpseo_metadesc"
    },
    "rankmath": {
        "title": "rank_math_title",
        "description": "rank_math_description"
    },
    "aioseo": {
        "title": "_aioseo_title",
        "description": "_aioseo_description"
    }
}

COLORS = {
    "bg":         "#0f1117",
    "panel":      "#1a1d27",
    "card":       "#22263a",
    "border":     "#2e3350",
    "accent":     "#4f8ef7",
    "accent2":    "#a78bfa",
    "success":    "#22d3a0",
    "error":      "#f87171",
    "warn":       "#fbbf24",
    "text":       "#e2e8f0",
    "text_dim":   "#8892a4",
    "input_bg":   "#181c2e",
    "hover":      "#2d3354",
}

# ─────────────────────────────────────────────
#  CONFIG MANAGER
# ─────────────────────────────────────────────
CONFIG_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "seo_sites_config.json")

def load_config():
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, "r") as f:
                return json.load(f)
        except:
            pass
    return {"sites": []}

def save_config(data):
    with open(CONFIG_FILE, "w") as f:
        json.dump(data, f, indent=2)

# ─────────────────────────────────────────────
#  WORDPRESS API
# ─────────────────────────────────────────────
def get_auth_header(username, app_password):
    creds = f"{username}:{app_password}"
    encoded = base64.b64encode(creds.encode()).decode()
    return {"Authorization": f"Basic {encoded}", "Content-Type": "application/json"}

def update_post_meta(site, post_id, post_type, meta_title, meta_description):
    base_url = site["url"].rstrip("/")
    plugin = site["seo_plugin"].lower()
    if plugin not in SEO_FIELDS:
        return False, f"Unknown plugin: {plugin}"
    fields = SEO_FIELDS[plugin]
    ep_map = {"post": "posts", "page": "pages", "product": "products"}
    endpoint = ep_map.get(post_type.lower(), post_type.lower() + "s")
    url = f"{base_url}/wp-json/wp/v2/{endpoint}/{post_id}"
    headers = get_auth_header(site["username"], site["app_password"])
    meta = {}
    if meta_title:
        meta[fields["title"]] = meta_title
    if meta_description:
        meta[fields["description"]] = meta_description
    if not meta:
        return False, "Both title and description are empty"
    try:
        r = requests.post(url, headers=headers, json={"meta": meta}, timeout=15)
        if r.status_code == 200:
            return True, f"OK ({plugin})"
        elif r.status_code == 401:
            return False, "Auth failed — check credentials"
        elif r.status_code == 403:
            return False, "Permission denied"
        elif r.status_code == 404:
            return False, f"Post {post_id} not found"
        else:
            try:
                msg = r.json().get("message", r.text[:120])
            except:
                msg = r.text[:120]
            return False, f"HTTP {r.status_code}: {msg}"
    except requests.exceptions.ConnectionError:
        return False, f"Cannot connect to {base_url}"
    except requests.exceptions.Timeout:
        return False, "Timeout"
    except Exception as e:
        return False, str(e)

def read_file(filepath):
    ext = os.path.splitext(filepath)[1].lower()
    rows = []
    if ext == ".csv":
        with open(filepath, newline='', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            for row in reader:
                rows.append({k.strip().lower(): (v.strip() if v else "") for k, v in row.items()})
    elif ext in [".xlsx", ".xls"]:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        headers = [str(cell.value).strip().lower() for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rd = {}
            for i, val in enumerate(row):
                if i < len(headers):
                    rd[headers[i]] = str(val).strip() if val is not None else ""
            if any(rd.values()):
                rows.append(rd)
    return rows


# ─────────────────────────────────────────────
#  MAIN APP
# ─────────────────────────────────────────────
class SEOUpdaterApp:
    def __init__(self, root):
        self.root = root
        self.root.title("WordPress SEO Meta Updater")
        self.root.geometry("1100x750")
        self.root.minsize(900, 650)
        self.root.configure(bg=COLORS["bg"])

        self.config_data = load_config()
        self.sites = self.config_data.get("sites", [])
        self.selected_file = tk.StringVar(value="No file selected")
        self.running = False

        self._apply_style()
        self._build_ui()
        self._refresh_sites_list()

    # ── STYLE ──
    def _apply_style(self):
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("TNotebook", background=COLORS["bg"], borderwidth=0)
        style.configure("TNotebook.Tab",
            background=COLORS["panel"], foreground=COLORS["text_dim"],
            padding=[18, 8], font=("Segoe UI", 10, "bold"), borderwidth=0)
        style.map("TNotebook.Tab",
            background=[("selected", COLORS["card"])],
            foreground=[("selected", COLORS["accent"])])
        style.configure("TFrame", background=COLORS["bg"])
        style.configure("Card.TFrame", background=COLORS["card"])
        style.configure("Panel.TFrame", background=COLORS["panel"])
        style.configure("TLabel",
            background=COLORS["bg"], foreground=COLORS["text"],
            font=("Segoe UI", 10))
        style.configure("Title.TLabel",
            background=COLORS["bg"], foreground=COLORS["text"],
            font=("Segoe UI", 18, "bold"))
        style.configure("Sub.TLabel",
            background=COLORS["bg"], foreground=COLORS["text_dim"],
            font=("Segoe UI", 9))
        style.configure("Card.TLabel",
            background=COLORS["card"], foreground=COLORS["text"],
            font=("Segoe UI", 10))
        style.configure("TEntry",
            fieldbackground=COLORS["input_bg"], foreground=COLORS["text"],
            bordercolor=COLORS["border"], relief="flat",
            insertcolor=COLORS["text"], font=("Segoe UI", 10))
        style.configure("TCombobox",
            fieldbackground=COLORS["input_bg"], foreground=COLORS["text"],
            selectbackground=COLORS["accent"], font=("Segoe UI", 10))
        style.configure("Treeview",
            background=COLORS["card"], foreground=COLORS["text"],
            fieldbackground=COLORS["card"], borderwidth=0,
            rowheight=34, font=("Segoe UI", 10))
        style.configure("Treeview.Heading",
            background=COLORS["panel"], foreground=COLORS["accent"],
            relief="flat", font=("Segoe UI", 10, "bold"))
        style.map("Treeview", background=[("selected", COLORS["hover"])])
        style.configure("TProgressbar",
            background=COLORS["accent"], troughcolor=COLORS["input_bg"],
            borderwidth=0, lightcolor=COLORS["accent"], darkcolor=COLORS["accent"])

    # ── BUTTON FACTORY ──
    def _btn(self, parent, text, command, color=None, width=None, small=False):
        bg = color or COLORS["accent"]
        fg = "#ffffff"
        pad_x = 14 if small else 20
        pad_y = 5 if small else 8
        fs = 9 if small else 10
        btn = tk.Button(parent, text=text, command=command,
            bg=bg, fg=fg, activebackground=COLORS["hover"],
            activeforeground=COLORS["text"], relief="flat", cursor="hand2",
            font=("Segoe UI", fs, "bold"), padx=pad_x, pady=pad_y,
            bd=0, highlightthickness=0)
        if width:
            btn.config(width=width)
        btn.bind("<Enter>", lambda e: btn.config(bg=COLORS["hover"]))
        btn.bind("<Leave>", lambda e: btn.config(bg=bg))
        return btn

    def _entry(self, parent, textvariable=None, show=None, width=30):
        e = tk.Entry(parent, textvariable=textvariable, show=show,
            bg=COLORS["input_bg"], fg=COLORS["text"],
            insertbackground=COLORS["text"], relief="flat",
            font=("Segoe UI", 10), width=width,
            highlightthickness=1, highlightbackground=COLORS["border"],
            highlightcolor=COLORS["accent"])
        return e

    # ── MAIN UI ──
    def _build_ui(self):
        # Header
        header = tk.Frame(self.root, bg=COLORS["panel"], height=64)
        header.pack(fill="x", side="top")
        header.pack_propagate(False)

        tk.Label(header, text="⚡", bg=COLORS["panel"],
            fg=COLORS["accent"], font=("Segoe UI", 20)).pack(side="left", padx=(20,6), pady=10)
        tk.Label(header, text="WordPress SEO Meta Updater",
            bg=COLORS["panel"], fg=COLORS["text"],
            font=("Segoe UI", 15, "bold")).pack(side="left", pady=10)
        tk.Label(header, text="Yoast · Rank Math · All in One SEO",
            bg=COLORS["panel"], fg=COLORS["text_dim"],
            font=("Segoe UI", 9)).pack(side="left", padx=14, pady=10)

        # Dependency warning
        if not HAS_REQUESTS:
            warn = tk.Frame(self.root, bg="#7c2d12")
            warn.pack(fill="x")
            tk.Label(warn, text="⚠  'requests' library not found. Run:  pip install requests openpyxl",
                bg="#7c2d12", fg="#fef3c7", font=("Segoe UI", 9, "bold")).pack(pady=4)

        # Notebook tabs
        nb = ttk.Notebook(self.root)
        nb.pack(fill="both", expand=True, padx=16, pady=12)

        self.tab_sites  = ttk.Frame(nb)
        self.tab_run    = ttk.Frame(nb)
        self.tab_log    = ttk.Frame(nb)

        nb.add(self.tab_sites, text="  🌐  Manage Sites  ")
        nb.add(self.tab_run,   text="  🚀  Run Updater  ")
        nb.add(self.tab_log,   text="  📋  Activity Log  ")

        self._build_sites_tab()
        self._build_run_tab()
        self._build_log_tab()

    # ─────────────────────────────────────────
    #  TAB 1 — MANAGE SITES
    # ─────────────────────────────────────────
    def _build_sites_tab(self):
        f = self.tab_sites
        f.configure(style="TFrame")

        left = tk.Frame(f, bg=COLORS["bg"], width=420)
        left.pack(side="left", fill="both", expand=False, padx=(0,8), pady=8)
        left.pack_propagate(False)

        right = tk.Frame(f, bg=COLORS["bg"])
        right.pack(side="left", fill="both", expand=True, pady=8)

        # ── ADD / EDIT FORM ──
        form_card = tk.Frame(left, bg=COLORS["card"],
            highlightthickness=1, highlightbackground=COLORS["border"])
        form_card.pack(fill="both", expand=True, padx=4)

        tk.Label(form_card, text="Add / Edit Site", bg=COLORS["card"],
            fg=COLORS["accent"], font=("Segoe UI", 12, "bold")).pack(anchor="w", padx=16, pady=(14,4))

        fields_frame = tk.Frame(form_card, bg=COLORS["card"])
        fields_frame.pack(fill="x", padx=16, pady=4)

        def lbl(text):
            return tk.Label(fields_frame, text=text, bg=COLORS["card"],
                fg=COLORS["text_dim"], font=("Segoe UI", 9))

        # Site Key
        lbl("Site Key (e.g. site1)").grid(row=0, column=0, sticky="w", pady=(6,1))
        self.v_key = tk.StringVar()
        self._entry(fields_frame, textvariable=self.v_key, width=34).grid(row=1, column=0, sticky="ew", pady=(0,6))

        # URL
        lbl("WordPress URL").grid(row=2, column=0, sticky="w", pady=(4,1))
        self.v_url = tk.StringVar()
        self._entry(fields_frame, textvariable=self.v_url, width=34).grid(row=3, column=0, sticky="ew", pady=(0,6))

        # Username
        lbl("Username").grid(row=4, column=0, sticky="w", pady=(4,1))
        self.v_user = tk.StringVar()
        self._entry(fields_frame, textvariable=self.v_user, width=34).grid(row=5, column=0, sticky="ew", pady=(0,6))

        # App Password
        lbl("Application Password").grid(row=6, column=0, sticky="w", pady=(4,1))
        self.v_pass = tk.StringVar()
        pass_frame = tk.Frame(fields_frame, bg=COLORS["card"])
        pass_frame.grid(row=7, column=0, sticky="ew", pady=(0,6))
        self.pass_entry = tk.Entry(pass_frame, textvariable=self.v_pass, show="●",
            bg=COLORS["input_bg"], fg=COLORS["text"], insertbackground=COLORS["text"],
            relief="flat", font=("Segoe UI", 10), width=28,
            highlightthickness=1, highlightbackground=COLORS["border"],
            highlightcolor=COLORS["accent"])
        self.pass_entry.pack(side="left", fill="x", expand=True)
        self.show_pass = False
        eye_btn = tk.Button(pass_frame, text="👁", command=self._toggle_pass,
            bg=COLORS["input_bg"], fg=COLORS["text_dim"], relief="flat",
            font=("Segoe UI", 10), cursor="hand2", bd=0, padx=6)
        eye_btn.pack(side="left")

        # SEO Plugin
        lbl("SEO Plugin").grid(row=8, column=0, sticky="w", pady=(4,1))
        self.v_plugin = tk.StringVar(value="yoast")
        plugin_frame = tk.Frame(fields_frame, bg=COLORS["card"])
        plugin_frame.grid(row=9, column=0, sticky="ew", pady=(0,10))
        for val, label in [("yoast","Yoast SEO"), ("rankmath","Rank Math"), ("aioseo","All in One SEO")]:
            rb = tk.Radiobutton(plugin_frame, text=label, variable=self.v_plugin, value=val,
                bg=COLORS["card"], fg=COLORS["text"], selectcolor=COLORS["input_bg"],
                activebackground=COLORS["card"], activeforeground=COLORS["accent"],
                font=("Segoe UI", 10), cursor="hand2")
            rb.pack(side="left", padx=(0,10))

        fields_frame.columnconfigure(0, weight=1)

        # Buttons
        btn_row = tk.Frame(form_card, bg=COLORS["card"])
        btn_row.pack(fill="x", padx=16, pady=(0,16))
        self._btn(btn_row, "💾  Save Site", self._save_site).pack(side="left", padx=(0,8))
        self._btn(btn_row, "🔍  Test Connection", self._test_connection,
            color=COLORS["card"]).pack(side="left", padx=(0,8))
        self._btn(btn_row, "✖  Clear", self._clear_form,
            color="#374151", small=True).pack(side="left")

        self.form_status = tk.Label(form_card, text="", bg=COLORS["card"],
            fg=COLORS["success"], font=("Segoe UI", 9))
        self.form_status.pack(anchor="w", padx=16, pady=(0,8))

        # ── SITES LIST ──
        tk.Label(right, text="Configured Sites", bg=COLORS["bg"],
            fg=COLORS["accent"], font=("Segoe UI", 12, "bold")).pack(anchor="w", pady=(4,6))

        tree_frame = tk.Frame(right, bg=COLORS["card"],
            highlightthickness=1, highlightbackground=COLORS["border"])
        tree_frame.pack(fill="both", expand=True)

        cols = ("key", "url", "plugin", "status")
        self.sites_tree = ttk.Treeview(tree_frame, columns=cols, show="headings", height=14)
        self.sites_tree.heading("key",    text="Site Key")
        self.sites_tree.heading("url",    text="URL")
        self.sites_tree.heading("plugin", text="SEO Plugin")
        self.sites_tree.heading("status", text="Status")
        self.sites_tree.column("key",    width=110, minwidth=80)
        self.sites_tree.column("url",    width=240, minwidth=160)
        self.sites_tree.column("plugin", width=120, minwidth=90)
        self.sites_tree.column("status", width=90,  minwidth=70)

        sb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.sites_tree.yview)
        self.sites_tree.configure(yscrollcommand=sb.set)
        self.sites_tree.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")
        self.sites_tree.bind("<Double-1>", self._load_site_to_form)

        action_row = tk.Frame(right, bg=COLORS["bg"])
        action_row.pack(fill="x", pady=(8,0))
        self._btn(action_row, "✏️  Edit Selected", self._load_site_to_form,
            color=COLORS["card"], small=True).pack(side="left", padx=(0,8))
        self._btn(action_row, "🗑️  Delete Selected", self._delete_site,
            color="#7f1d1d", small=True).pack(side="left")
        tk.Label(action_row, text="Double-click a row to edit",
            bg=COLORS["bg"], fg=COLORS["text_dim"],
            font=("Segoe UI", 8)).pack(side="right")

    def _toggle_pass(self):
        self.show_pass = not self.show_pass
        self.pass_entry.config(show="" if self.show_pass else "●")

    def _clear_form(self):
        self.v_key.set(""); self.v_url.set("")
        self.v_user.set(""); self.v_pass.set("")
        self.v_plugin.set("yoast")
        self.form_status.config(text="")
        self._editing_index = None

    def _save_site(self):
        key    = self.v_key.get().strip()
        url    = self.v_url.get().strip()
        user   = self.v_user.get().strip()
        pwd    = self.v_pass.get().strip()
        plugin = self.v_plugin.get()

        if not all([key, url, user, pwd]):
            self.form_status.config(text="⚠ All fields are required", fg=COLORS["warn"])
            return
        if not url.startswith("http"):
            self.form_status.config(text="⚠ URL must start with https://", fg=COLORS["warn"])
            return

        new_site = {"key": key, "url": url, "username": user,
                    "app_password": pwd, "seo_plugin": plugin}

        # Check if editing existing
        idx = getattr(self, "_editing_index", None)
        if idx is not None:
            self.sites[idx] = new_site
            self.form_status.config(text=f"✓ Updated '{key}'", fg=COLORS["success"])
        else:
            # Check duplicate key
            if any(s["key"] == key for s in self.sites):
                self.form_status.config(text=f"⚠ Key '{key}' already exists", fg=COLORS["warn"])
                return
            self.sites.append(new_site)
            self.form_status.config(text=f"✓ Added '{key}'", fg=COLORS["success"])

        self.config_data["sites"] = self.sites
        save_config(self.config_data)
        self._refresh_sites_list()
        self._clear_form()

    def _delete_site(self):
        sel = self.sites_tree.selection()
        if not sel:
            messagebox.showwarning("No Selection", "Please select a site to delete.")
            return
        item = self.sites_tree.item(sel[0])
        key = item["values"][0]
        if messagebox.askyesno("Confirm Delete", f"Delete site '{key}'?"):
            self.sites = [s for s in self.sites if s["key"] != key]
            self.config_data["sites"] = self.sites
            save_config(self.config_data)
            self._refresh_sites_list()

    def _load_site_to_form(self, event=None):
        sel = self.sites_tree.selection()
        if not sel:
            return
        item = self.sites_tree.item(sel[0])
        key = item["values"][0]
        for i, s in enumerate(self.sites):
            if s["key"] == key:
                self.v_key.set(s["key"])
                self.v_url.set(s["url"])
                self.v_user.set(s["username"])
                self.v_pass.set(s["app_password"])
                self.v_plugin.set(s["seo_plugin"])
                self._editing_index = i
                self.form_status.config(
                    text=f"Editing '{key}' — click Save to update", fg=COLORS["accent2"])
                break

    def _test_connection(self):
        url  = self.v_url.get().strip()
        user = self.v_user.get().strip()
        pwd  = self.v_pass.get().strip()
        if not all([url, user, pwd]):
            self.form_status.config(text="⚠ Enter URL, username & password first", fg=COLORS["warn"])
            return
        self.form_status.config(text="⏳ Testing...", fg=COLORS["text_dim"])
        self.root.update()
        try:
            headers = get_auth_header(user, pwd)
            r = requests.get(f"{url.rstrip('/')}/wp-json/wp/v2/users/me",
                             headers=headers, timeout=10)
            if r.status_code == 200:
                name = r.json().get("name", "Unknown")
                self.form_status.config(
                    text=f"✓ Connected! Logged in as: {name}", fg=COLORS["success"])
            elif r.status_code == 401:
                self.form_status.config(text="✗ Auth failed — check credentials", fg=COLORS["error"])
            else:
                self.form_status.config(
                    text=f"✗ HTTP {r.status_code}", fg=COLORS["error"])
        except Exception as e:
            self.form_status.config(text=f"✗ {str(e)[:60]}", fg=COLORS["error"])

    def _refresh_sites_list(self):
        for item in self.sites_tree.get_children():
            self.sites_tree.delete(item)
        plugin_labels = {"yoast": "Yoast SEO", "rankmath": "Rank Math", "aioseo": "All in One SEO"}
        for s in self.sites:
            self.sites_tree.insert("", "end", values=(
                s["key"], s["url"],
                plugin_labels.get(s["seo_plugin"], s["seo_plugin"]),
                "✓ Saved"
            ))

    # ─────────────────────────────────────────
    #  TAB 2 — RUN UPDATER
    # ─────────────────────────────────────────
    def _build_run_tab(self):
        f = self.tab_run
        f.configure(style="TFrame")

        # Top controls
        top = tk.Frame(f, bg=COLORS["card"],
            highlightthickness=1, highlightbackground=COLORS["border"])
        top.pack(fill="x", padx=4, pady=(8,0))

        tk.Label(top, text="Select Input File (CSV or Excel)",
            bg=COLORS["card"], fg=COLORS["accent"],
            font=("Segoe UI", 12, "bold")).pack(anchor="w", padx=16, pady=(14,4))

        file_row = tk.Frame(top, bg=COLORS["card"])
        file_row.pack(fill="x", padx=16, pady=(0,12))

        self._btn(file_row, "📂  Browse File", self._browse_file,
            color="#1d4ed8").pack(side="left", padx=(0,12))

        tk.Label(file_row, textvariable=self.selected_file,
            bg=COLORS["card"], fg=COLORS["text_dim"],
            font=("Segoe UI", 9), wraplength=500).pack(side="left")

        # CSV format hint
        hint = tk.Frame(f, bg=COLORS["panel"],
            highlightthickness=1, highlightbackground=COLORS["border"])
        hint.pack(fill="x", padx=4, pady=(8,0))
        tk.Label(hint, text="📌  CSV Required Columns:",
            bg=COLORS["panel"], fg=COLORS["accent2"],
            font=("Segoe UI", 9, "bold")).pack(side="left", padx=(12,6), pady=6)
        tk.Label(hint,
            text="site_key  |  post_id  |  post_type  |  meta_title  |  meta_description",
            bg=COLORS["panel"], fg=COLORS["text_dim"],
            font=("Courier", 9)).pack(side="left", pady=6)

        # Progress area
        prog_card = tk.Frame(f, bg=COLORS["card"],
            highlightthickness=1, highlightbackground=COLORS["border"])
        prog_card.pack(fill="x", padx=4, pady=8)

        prog_top = tk.Frame(prog_card, bg=COLORS["card"])
        prog_top.pack(fill="x", padx=16, pady=(12,6))
        tk.Label(prog_top, text="Progress", bg=COLORS["card"],
            fg=COLORS["text"], font=("Segoe UI", 10, "bold")).pack(side="left")
        self.progress_label = tk.Label(prog_top, text="",
            bg=COLORS["card"], fg=COLORS["text_dim"], font=("Segoe UI", 9))
        self.progress_label.pack(side="right")

        self.progress_var = tk.DoubleVar()
        self.progressbar = ttk.Progressbar(prog_card, variable=self.progress_var,
            maximum=100, style="TProgressbar")
        self.progressbar.pack(fill="x", padx=16, pady=(0,10))

        # Stats row
        self.stats_frame = tk.Frame(prog_card, bg=COLORS["card"])
        self.stats_frame.pack(fill="x", padx=16, pady=(0,12))

        self.stat_total   = self._stat_box(self.stats_frame, "Total",     "0", COLORS["text_dim"])
        self.stat_success = self._stat_box(self.stats_frame, "Success",   "0", COLORS["success"])
        self.stat_fail    = self._stat_box(self.stats_frame, "Failed",    "0", COLORS["error"])
        self.stat_skip    = self._stat_box(self.stats_frame, "Skipped",   "0", COLORS["warn"])

        # Run / Stop
        btn_row = tk.Frame(f, bg=COLORS["bg"])
        btn_row.pack(fill="x", padx=4, pady=4)
        self.run_btn  = self._btn(btn_row, "▶  Start Update", self._start_update, color="#15803d")
        self.run_btn.pack(side="left", padx=(0,10))
        self.stop_btn = self._btn(btn_row, "⏹  Stop", self._stop_update, color="#7f1d1d")
        self.stop_btn.pack(side="left")
        self.stop_btn.config(state="disabled")

        # Live feed
        tk.Label(f, text="Live Feed", bg=COLORS["bg"],
            fg=COLORS["text_dim"], font=("Segoe UI", 9, "bold")).pack(anchor="w", padx=6, pady=(6,2))
        self.live_text = scrolledtext.ScrolledText(f,
            bg=COLORS["input_bg"], fg=COLORS["text"],
            font=("Consolas", 9), relief="flat", height=10,
            insertbackground=COLORS["text"],
            highlightthickness=1, highlightbackground=COLORS["border"])
        self.live_text.pack(fill="both", expand=True, padx=4, pady=(0,8))
        self.live_text.config(state="disabled")

    def _stat_box(self, parent, label, value, color):
        box = tk.Frame(parent, bg=COLORS["panel"], padx=16, pady=8)
        box.pack(side="left", padx=(0,8))
        v_var = tk.StringVar(value=value)
        tk.Label(box, textvariable=v_var, bg=COLORS["panel"],
            fg=color, font=("Segoe UI", 18, "bold")).pack()
        tk.Label(box, text=label, bg=COLORS["panel"],
            fg=COLORS["text_dim"], font=("Segoe UI", 8)).pack()
        return v_var

    def _browse_file(self):
        path = filedialog.askopenfilename(
            title="Select CSV or Excel file",
            filetypes=[("CSV/Excel", "*.csv *.xlsx *.xls"), ("All files", "*.*")])
        if path:
            self.selected_file.set(path)

    def _live_log(self, msg, color=None):
        self.live_text.config(state="normal")
        tag = f"color_{id(msg)}"
        self.live_text.tag_config(tag, foreground=color or COLORS["text"])
        self.live_text.insert("end", msg + "\n", tag)
        self.live_text.see("end")
        self.live_text.config(state="disabled")

    def _start_update(self):
        if not self.sites:
            messagebox.showwarning("No Sites", "Please add at least one site in the 'Manage Sites' tab.")
            return
        filepath = self.selected_file.get()
        if filepath == "No file selected" or not os.path.exists(filepath):
            messagebox.showwarning("No File", "Please select a valid CSV or Excel file.")
            return
        if not HAS_REQUESTS:
            messagebox.showerror("Missing Library",
                "The 'requests' library is not installed.\nRun: pip install requests openpyxl")
            return

        self.running = True
        self.run_btn.config(state="disabled")
        self.stop_btn.config(state="normal")
        self.live_text.config(state="normal")
        self.live_text.delete("1.0", "end")
        self.live_text.config(state="disabled")
        self.progress_var.set(0)

        thread = threading.Thread(target=self._run_update, args=(filepath,), daemon=True)
        thread.start()

    def _stop_update(self):
        self.running = False
        self._live_log("⏹  Stopped by user.", COLORS["warn"])
        self.run_btn.config(state="normal")
        self.stop_btn.config(state="disabled")

    def _run_update(self, filepath):
        sites_map = {s["key"]: s for s in self.sites}
        try:
            rows = read_file(filepath)
        except Exception as e:
            self.root.after(0, self._live_log, f"✗ Error reading file: {e}", COLORS["error"])
            self.root.after(0, self.run_btn.config, {"state": "normal"})
            self.root.after(0, self.stop_btn.config, {"state": "disabled"})
            return

        total = len(rows)
        success = fail = skip = 0

        self.root.after(0, self._live_log,
            f"📂  Loaded {total} rows from file", COLORS["accent"])
        self.root.after(0, self._live_log,
            f"{'─'*55}", COLORS["border"])
        self.root.after(0, self.stat_total.set, str(total))

        for i, row in enumerate(rows, 1):
            if not self.running:
                break

            site_key  = row.get("site_key", "").strip()
            post_id   = row.get("post_id", "").strip()
            post_type = row.get("post_type", "post").strip() or "post"
            meta_t    = row.get("meta_title", "").strip()
            meta_d    = row.get("meta_description", "").strip()

            pct = int((i / total) * 100)
            self.root.after(0, self.progress_var.set, pct)
            self.root.after(0, self.progress_label.config,
                {"text": f"{i} / {total}  ({pct}%)"})

            prefix = f"[{i}/{total}]"

            if site_key not in sites_map:
                skip += 1
                self.root.after(0, self._live_log,
                    f"{prefix}  ⚠  Unknown site '{site_key}' — skipped", COLORS["warn"])
                self.root.after(0, self.stat_skip.set, str(skip))
                continue

            if not post_id.isdigit():
                skip += 1
                self.root.after(0, self._live_log,
                    f"{prefix}  ⚠  Invalid post_id '{post_id}' — skipped", COLORS["warn"])
                self.root.after(0, self.stat_skip.set, str(skip))
                continue

            ok, msg = update_post_meta(sites_map[site_key], post_id, post_type, meta_t, meta_d)
            if ok:
                success += 1
                self.root.after(0, self._live_log,
                    f"{prefix}  ✓  {site_key} / post {post_id} — {msg}", COLORS["success"])
                self.root.after(0, self.stat_success.set, str(success))
            else:
                fail += 1
                self.root.after(0, self._live_log,
                    f"{prefix}  ✗  {site_key} / post {post_id} — {msg}", COLORS["error"])
                self.root.after(0, self.stat_fail.set, str(fail))

            time.sleep(0.4)

        self.root.after(0, self._live_log, f"{'─'*55}", COLORS["border"])
        self.root.after(0, self._live_log,
            f"✅  Done!  Total: {total}  ✓ {success}  ✗ {fail}  ⚠ {skip}", COLORS["accent2"])
        self.root.after(0, self.progress_var.set, 100)
        self.root.after(0, self.run_btn.config, {"state": "normal"})
        self.root.after(0, self.stop_btn.config, {"state": "disabled"})
        self.running = False

    # ─────────────────────────────────────────
    #  TAB 3 — LOG
    # ─────────────────────────────────────────
    def _build_log_tab(self):
        f = self.tab_log
        f.configure(style="TFrame")

        tk.Label(f, text="Activity Log — all updates are recorded here",
            bg=COLORS["bg"], fg=COLORS["text_dim"],
            font=("Segoe UI", 9)).pack(anchor="w", padx=8, pady=(8,4))

        self.log_text = scrolledtext.ScrolledText(f,
            bg=COLORS["input_bg"], fg=COLORS["text"],
            font=("Consolas", 9), relief="flat",
            insertbackground=COLORS["text"],
            highlightthickness=1, highlightbackground=COLORS["border"])
        self.log_text.pack(fill="both", expand=True, padx=4)

        btn_row = tk.Frame(f, bg=COLORS["bg"])
        btn_row.pack(fill="x", padx=4, pady=8)
        self._btn(btn_row, "💾  Save Log to File", self._save_log, color="#1d4ed8", small=True).pack(side="left", padx=(0,8))
        self._btn(btn_row, "🗑  Clear Log", self._clear_log, color="#374151", small=True).pack(side="left")

        # Redirect live feed to log tab too
        self.log_lines = []

    def _append_log(self, msg):
        ts = datetime.now().strftime("%H:%M:%S")
        line = f"[{ts}]  {msg}\n"
        self.log_text.config(state="normal")
        self.log_text.insert("end", line)
        self.log_text.see("end")
        self.log_text.config(state="disabled")
        self.log_lines.append(line)

    def _save_log(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("Text file", "*.txt")],
            title="Save log as")
        if path:
            with open(path, "w", encoding="utf-8") as f:
                f.write("".join(self.log_lines))
            messagebox.showinfo("Saved", f"Log saved to:\n{path}")

    def _clear_log(self):
        self.log_text.config(state="normal")
        self.log_text.delete("1.0", "end")
        self.log_text.config(state="disabled")
        self.log_lines.clear()


# ─────────────────────────────────────────────
#  ENTRY POINT
# ─────────────────────────────────────────────
if __name__ == "__main__":
    root = tk.Tk()
    root.tk_setPalette(background=COLORS["bg"], foreground=COLORS["text"])
    try:
        root.iconbitmap("")
    except:
        pass
    app = SEOUpdaterApp(root)
    root.mainloop()
