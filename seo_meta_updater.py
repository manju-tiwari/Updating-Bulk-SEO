#!/usr/bin/env python3
"""
=============================================================
  Multi-Site WordPress SEO Meta Updater
  Works with: Yoast SEO | Rank Math | All in One SEO
  Input: CSV or Excel file
=============================================================
"""

import csv
import json
import base64
import time
import sys
import os
from datetime import datetime

# Try to import optional libraries
try:
    import requests
    HAS_REQUESTS = True
except ImportError:
    HAS_REQUESTS = False

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ─────────────────────────────────────────────
#  CONFIGURATION — Edit this section
# ─────────────────────────────────────────────

SITES = {
    "site1": {
        "url": "https://yoursite1.com",
        "username": "admin",
        "app_password": "xxxx xxxx xxxx xxxx xxxx xxxx",   # WordPress App Password
        "seo_plugin": "yoast"       # Options: yoast | rankmath | aioseo
    },
    "site2": {
        "url": "https://yoursite2.com",
        "username": "admin",
        "app_password": "xxxx xxxx xxxx xxxx xxxx xxxx",
        "seo_plugin": "rankmath"
    },
    "site3": {
        "url": "https://yoursite3.com",
        "username": "admin",
        "app_password": "xxxx xxxx xxxx xxxx xxxx xxxx",
        "seo_plugin": "aioseo"
    },
    # Add more sites here...
}

# Input file path (CSV or XLSX)
INPUT_FILE = "seo_updates.csv"   # Change to your file name

# Delay between API calls (seconds) — avoids rate limiting
REQUEST_DELAY = 0.5

# Log file
LOG_FILE = f"seo_update_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"

# ─────────────────────────────────────────────
#  SEO PLUGIN FIELD MAPPINGS
# ─────────────────────────────────────────────

SEO_FIELDS = {
    "yoast": {
        "title": "_yoast_wpseo_title",
        "description": "_yoast_wpseo_metadesc"
    },
    "rankmath": {
        "title": "rank_math_title",
        "description": "rank_math_description"
    },
    "aioseo": {
        "title": "_aioseo_title",
        "description": "_aioseo_description"
    }
}

# ─────────────────────────────────────────────
#  LOGGER
# ─────────────────────────────────────────────

log_lines = []

def log(message, level="INFO"):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{timestamp}] [{level}] {message}"
    print(line)
    log_lines.append(line)

def save_log():
    with open(LOG_FILE, "w", encoding="utf-8") as f:
        f.write("\n".join(log_lines))
    log(f"Log saved to: {LOG_FILE}")

# ─────────────────────────────────────────────
#  FILE READER (CSV + XLSX)
# ─────────────────────────────────────────────

def read_input_file(filepath):
    """
    Reads CSV or Excel file.
    Expected columns:
      site_key | post_id | post_type | meta_title | meta_description
    """
    ext = os.path.splitext(filepath)[1].lower()
    rows = []

    if ext == ".csv":
        with open(filepath, newline='', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            for row in reader:
                rows.append({k.strip().lower(): v.strip() for k, v in row.items()})

    elif ext in [".xlsx", ".xls"]:
        if not HAS_OPENPYXL:
            log("openpyxl not installed. Run: pip install openpyxl", "ERROR")
            sys.exit(1)
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        headers = [str(cell.value).strip().lower() for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_data = {}
            for i, val in enumerate(row):
                if i < len(headers):
                    row_data[headers[i]] = str(val).strip() if val is not None else ""
            if any(row_data.values()):
                rows.append(row_data)
    else:
        log(f"Unsupported file type: {ext}. Use .csv or .xlsx", "ERROR")
        sys.exit(1)

    log(f"Loaded {len(rows)} rows from {filepath}")
    return rows

# ─────────────────────────────────────────────
#  WORDPRESS API FUNCTIONS
# ─────────────────────────────────────────────

def get_auth_header(username, app_password):
    credentials = f"{username}:{app_password}"
    encoded = base64.b64encode(credentials.encode()).decode()
    return {"Authorization": f"Basic {encoded}", "Content-Type": "application/json"}

def detect_post_type_endpoint(post_type):
    """Map post type to WP REST API endpoint."""
    mapping = {
        "post": "posts",
        "page": "pages",
        "product": "products",   # WooCommerce
    }
    return mapping.get(post_type.lower(), post_type.lower() + "s")

def update_post_meta(site_config, post_id, post_type, meta_title, meta_description):
    """
    Updates SEO meta fields via WordPress REST API.
    Returns: (success: bool, message: str)
    """
    base_url = site_config["url"].rstrip("/")
    username = site_config["username"]
    app_password = site_config["app_password"]
    plugin = site_config["seo_plugin"].lower()

    if plugin not in SEO_FIELDS:
        return False, f"Unknown SEO plugin: {plugin}. Use: yoast, rankmath, aioseo"

    fields = SEO_FIELDS[plugin]
    endpoint = detect_post_type_endpoint(post_type)
    url = f"{base_url}/wp-json/wp/v2/{endpoint}/{post_id}"
    headers = get_auth_header(username, app_password)

    # Build meta payload — only include non-empty values
    meta_payload = {}
    if meta_title:
        meta_payload[fields["title"]] = meta_title
    if meta_description:
        meta_payload[fields["description"]] = meta_description

    if not meta_payload:
        return False, "Both meta_title and meta_description are empty — skipped"

    payload = {"meta": meta_payload}

    try:
        response = requests.post(url, headers=headers, json=payload, timeout=15)

        if response.status_code == 200:
            return True, f"Updated successfully (Plugin: {plugin})"
        elif response.status_code == 401:
            return False, "Authentication failed — check username/app_password"
        elif response.status_code == 403:
            return False, "Permission denied — user may not have edit rights"
        elif response.status_code == 404:
            return False, f"Post ID {post_id} not found on {base_url}"
        else:
            detail = response.json().get("message", response.text[:200])
            return False, f"HTTP {response.status_code}: {detail}"

    except requests.exceptions.ConnectionError:
        return False, f"Cannot connect to {base_url} — check URL"
    except requests.exceptions.Timeout:
        return False, "Request timed out"
    except Exception as e:
        return False, f"Unexpected error: {str(e)}"

# ─────────────────────────────────────────────
#  MAIN RUNNER
# ─────────────────────────────────────────────

def run():
    if not HAS_REQUESTS:
        print("ERROR: 'requests' library not installed.")
        print("Run: pip install requests openpyxl")
        sys.exit(1)

    log("=" * 60)
    log("  Multi-Site WordPress SEO Meta Updater Started")
    log("=" * 60)

    # Check input file
    if not os.path.exists(INPUT_FILE):
        log(f"Input file not found: {INPUT_FILE}", "ERROR")
        log("Please create your CSV/Excel file and update INPUT_FILE path.")
        save_log()
        sys.exit(1)

    # Read input
    rows = read_input_file(INPUT_FILE)

    # Stats
    total = len(rows)
    success_count = 0
    fail_count = 0
    skip_count = 0

    # Process each row
    for i, row in enumerate(rows, 1):
        site_key   = row.get("site_key", "").strip()
        post_id    = row.get("post_id", "").strip()
        post_type  = row.get("post_type", "post").strip() or "post"
        meta_title = row.get("meta_title", "").strip()
        meta_desc  = row.get("meta_description", "").strip()

        log(f"\n[{i}/{total}] Processing: site={site_key} | post_id={post_id}")

        # Validate site key
        if site_key not in SITES:
            log(f"  ⚠ Unknown site_key '{site_key}' — skipping", "WARN")
            skip_count += 1
            continue

        # Validate post ID
        if not post_id.isdigit():
            log(f"  ⚠ Invalid post_id '{post_id}' — skipping", "WARN")
            skip_count += 1
            continue

        # Update meta
        site_config = SITES[site_key]
        success, message = update_post_meta(
            site_config, post_id, post_type, meta_title, meta_desc
        )

        if success:
            log(f"  ✓ {message}")
            success_count += 1
        else:
            log(f"  ✗ {message}", "ERROR")
            fail_count += 1

        # Rate limiting
        time.sleep(REQUEST_DELAY)

    # Summary
    log("\n" + "=" * 60)
    log("  SUMMARY")
    log("=" * 60)
    log(f"  Total Rows   : {total}")
    log(f"  ✓ Successful : {success_count}")
    log(f"  ✗ Failed     : {fail_count}")
    log(f"  ⚠ Skipped    : {skip_count}")
    log("=" * 60)

    save_log()

if __name__ == "__main__":
    run()
